import pdfplumber
import openpyxl
import re
import logging
from datetime import datetime, date as date_cls
from categories import normalize_description

logger = logging.getLogger(__name__)


def process_pdf(file_path: str) -> tuple[list[dict], float | None, str]:
    """Parse a Hapoalim עו"ש PDF.

    Signature kept as a 3-tuple for backward compatibility with bot.py.
    The third element is always "poalim" or "unknown" — Cal PDFs are no
    longer supported (Cal is provided as Excel now).
    """
    try:
        with pdfplumber.open(file_path) as pdf:
            full_text = ""
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    full_text += text + "\n"
    except Exception as e:
        logger.error(f"PDF read error: {e}")
        return [], None, "unknown"

    transactions, closing_balance = _parse_poalim_evosh(full_text)
    if not transactions:
        return [], None, "unknown"
    logger.info(f"פועלים: נמצאו {len(transactions)} עסקאות, יתרה={closing_balance}")
    return transactions, closing_balance, "poalim"


def _fix_rtl(text: str) -> str:
    """pdfplumber reverses Hebrew text two ways: word order and each
    word's letters. Fix: reverse word order, then reverse the letters
    of words that contain Hebrew. Latin words (WOLT, MBD, aliexpress)
    stay as-is — otherwise we get "TLOW", "DBM", "sserpxeila"."""
    def is_hebrew(w: str) -> bool:
        return any('֐' <= c <= '׿' for c in w)
    return " ".join(
        (w[::-1] if is_hebrew(w) else w)
        for w in text.split()[::-1]
    )


def _parse_poalim_evosh(text: str) -> list[dict]:
    """
    Poalim bank statement (עו"ש) PDF.
    pdfplumber extracts each row as TWO lines:
      Line 1: ##
      Line 2: ₪BALANCE  AMOUNT  DESCRIPTION(reversed)  DATE

    Poalim is the main account — money sits here and everything is debited
    from it. We classify each transaction:
      - income       : credits (salary, refunds, transfers in)
      - max_summary  : the monthly Max credit-card debit (do NOT sum with
                       the Max file rows — that would double-count)
      - expense      : everything else
    """
    # The "##" marker appears either on its own line ("##\n₪...") or — for
    # the very first (most recent) transaction — inline with the data
    # ("## ₪..."). Allow either with \s+ so we don't drop the latest row.
    pattern = re.compile(
        r'##\s+(₪[\d,]+\.?\d*)\s+([\d,]+\.?\d*)\s+(.+?)\s+(\d{2}/\d{2}/\d{4})'
    )

    # Hebrew keywords (reversed, since pdfplumber returns text reversed)
    # that indicate INCOME — credits to the account.
    # NOTE: "העברה" is intentionally NOT here — it matches outgoing transfers
    # too (e.g. "העברה לאחר נייד") and would mis-classify expenses as income.
    INCOME_KEYWORDS = [
        "תרוכשמ",        # משכורת
        'ת"פומ',          # מופ"ת
        "יוכיז",          # זיכוי
        "קנעמ",           # מענק
        "םירמושה תצובק",  # קבוצת השומרים
        "רזחה",           # החזר
        "תיביר",          # ריבית
    ]

    # Identifies the monthly Max credit-card debit line in the Poalim
    # statement. We require the full distinctive phrase "מקס איט" (reversed:
    # "טיא סקמ") — matching just "מקס"/"סקמ" caught dozens of unrelated rows
    # (any word containing those 3 letters: מקסים, מקסי, etc).
    MAX_SUMMARY_PATTERNS = [
        "טיא סקמ",        # מקס איט   (the Max debit line in Hapoalim)
        "יסנניפ טיא סקמ", # מקס איט פיננסי
    ]
    # Cal credit card monthly debit line. The description in the PDF after
    # RTL fix is exactly "כאל". Reversed in the raw text: "לאכ".
    CAL_SUMMARY_DESCRIPTIONS = {"כאל"}

    transactions = []
    balances_by_date: list[tuple[datetime, float]] = []

    for m in pattern.finditer(text):
        balance_str = m.group(1)
        amount_str = m.group(2)
        desc_raw = m.group(3).strip()
        date = m.group(4)
        description = _fix_rtl(desc_raw)

        try:
            amount = float(amount_str.replace(",", ""))
            if amount <= 0 or amount > 200000:
                continue
        except ValueError:
            continue

        try:
            balance = float(balance_str.replace("₪", "").replace(",", ""))
            dt = datetime.strptime(date, "%d/%m/%Y")
            balances_by_date.append((dt, balance))
        except ValueError:
            pass

        if any(kw in desc_raw for kw in INCOME_KEYWORDS):
            tx_type = "income"
        elif any(kw in desc_raw for kw in MAX_SUMMARY_PATTERNS):
            tx_type = "max_summary"
        elif description in CAL_SUMMARY_DESCRIPTIONS:
            tx_type = "cal_summary"
        else:
            tx_type = "expense"

        transactions.append({
            "date": date,
            "description": normalize_description(description),
            "amount": amount,
            "source": 'פועלים עו"ש',
            "type": tx_type,
        })

    closing_balance: float | None = None
    if balances_by_date:
        balances_by_date.sort(key=lambda x: x[0], reverse=True)
        closing_balance = balances_by_date[0][1]

    return transactions, closing_balance


def _norm_cell(c) -> str:
    """Normalise a cell value for header matching: strip and replace embedded
    newlines so 'תאריך\\nעסקה' matches 'תאריך עסקה'."""
    if c is None:
        return ""
    return str(c).replace("\n", " ").replace("\r", " ").strip()


def _format_excel_date(v) -> str:
    """Normalise Excel date cells to DD/MM/YYYY.

    Cal exports dates as datetime objects (e.g. 2026-05-07 00:00:00).
    Max exports them as DD-MM-YYYY strings. Pass-through anything else
    that already looks like a date.
    """
    if isinstance(v, datetime):
        return v.strftime("%d/%m/%Y")
    if isinstance(v, date_cls):
        return v.strftime("%d/%m/%Y")
    s = str(v).strip()
    # Match "DD-MM-YYYY" or "DD/MM/YYYY" or "DD.MM.YYYY"
    m = re.match(r"^(\d{1,2})[-/.](\d{1,2})[-/.](\d{2,4})$", s)
    if m:
        d, mo, y = m.groups()
        if len(y) == 2:
            y = "20" + y
        return f"{d.zfill(2)}/{mo.zfill(2)}/{y}"
    # Match "YYYY-MM-DD..." (ISO-ish)
    m = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})", s)
    if m:
        y, mo, d = m.groups()
        return f"{d.zfill(2)}/{mo.zfill(2)}/{y}"
    return s


def _is_cal_workbook(wb) -> bool:
    """Detect a Cal-Hapoalim credit-card Excel export.

    Cal sheets are named after the account ("הפועלים <account-number>")
    and the title row mentions "כרטיס מאסטרקארד" / "פירוט עסקאות לחשבון
    הפועלים". Max exports name the sheet "עסקאות במועד החיוב" instead.
    """
    if any("הפועלים" in (sn or "") for sn in wb.sheetnames):
        return True
    # Fall back to looking at row 1 of the first sheet.
    try:
        ws = wb[wb.sheetnames[0]]
        row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        title = " ".join(_norm_cell(c) for c in row1)
        if "מאסטרקארד" in title or ("הפועלים" in title and "פירוט" in title):
            return True
    except Exception:
        pass
    return False


def process_excel(file_path: str) -> list[dict]:
    """Parse a credit-card Excel — auto-detects Cal vs Max.

    Returns a list of transaction dicts. Each has a 'source' field of
    either 'כאל פועלים' (Cal) or 'מקס' (Max), which is how the caller
    routes the file into the right bucket.
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        logger.error(f"Excel read error: {e}")
        return []

    if _is_cal_workbook(wb):
        transactions = _parse_cal_excel(wb)
        logger.info(f"כאל: נמצאו {len(transactions)} עסקאות")
    else:
        transactions = _parse_max_excel(wb)
        logger.info(f"מקס: נמצאו {len(transactions)} עסקאות")
    return transactions


def _parse_cal_excel(wb) -> list[dict]:
    """Parse a Cal credit-card Excel export.

    Layout (sheet name like 'הפועלים 732-XXXXXX'):
      R1 title, R2 blank, R3 "עסקאות לחיוב ב-DD/MM/YYYY: AMOUNT ₪",
      R4 "עסקאות בחיוב מיידי AMOUNT ₪", R5 header, R6+ data,
      footer line starts with "את המידע המלא".

    Header columns (R5):
      A=תאריך עסקה, B=שם בית עסק, C=סכום עסקה, D=סכום חיוב,
      E=סוג עסקה, F=ענף, G=הערות

    We use column D (סכום חיוב = actually-charged amount) so installments
    contribute their per-instalment amount, not the gross transaction.
    Merchant names are already in correct reading order in Excel — do
    NOT pass them through _fix_rtl.
    """
    transactions: list[dict] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Find header row by content
        header_row = None
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            cells = [_norm_cell(c) for c in row]
            if any("תאריך" in c and "עסקה" in c for c in cells) and \
               any("שם בית" in c for c in cells):
                header_row = i
                break
        if header_row is None:
            continue

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or len(row) < 4:
                continue
            date_val, desc_val, _orig_amt, charge_val = row[0], row[1], row[2], row[3]
            note_val = row[6] if len(row) > 6 else ""

            # Skip footer lines and blanks
            if date_val is None and desc_val is None and charge_val is None:
                continue
            desc = _norm_cell(desc_val)
            if not desc or "את המידע" in desc:
                continue

            try:
                amount = float(str(charge_val).replace(",", "").replace("₪", "").strip())
            except (ValueError, TypeError, AttributeError):
                continue
            # Keep negatives — they are refunds and Hapoalim's monthly Cal
            # debit already nets them out, so dropping them would break the
            # self-verification check.
            if amount == 0 or abs(amount) > 200000:
                continue

            note = _norm_cell(note_val)
            installment_match = re.match(r"תשלום (\d+) מתוך (\d+)", note)
            if installment_match:
                desc += f" (תשלום {installment_match.group(1)}/{installment_match.group(2)})"

            transactions.append({
                "date": _format_excel_date(date_val),
                "description": normalize_description(desc),
                "amount": amount,
                "source": "כאל פועלים",
                "type": "expense",
            })

    return transactions


def _parse_max_excel(wb) -> list[dict]:
    """Parse a Max credit-card Excel export.

    Max files have up to two sheets:
      - 'עסקאות במועד החיוב'      (regular monthly billing)
      - 'עסקאות חו"ל ומט"ח'        (foreign / immediate-debit)

    Each sheet has the same column structure:
      Headers at row 4 — A=תאריך עסקה, B=שם בית העסק, F=סכום חיוב,
      J=תאריך חיוב. The row after the last transaction is blank, then
      'סך הכל' / total. Parse both sheets; ignore the footer.
    """
    transactions: list[dict] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Locate header row and column indices
        header_row = None
        col_map: dict[str, int] = {}
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            cells = [_norm_cell(c) for c in row]
            if any("תאריך" in c and "עסקה" in c for c in cells) and \
               any("שם בית" in c or "תיאור" in c for c in cells):
                header_row = i
                for j, cell in enumerate(cells):
                    if "תאריך" in cell and "עסקה" in cell and "date" not in col_map:
                        col_map["date"] = j
                    elif ("שם בית" in cell or "תיאור" in cell) and "description" not in col_map:
                        col_map["description"] = j
                    elif "סכום חיוב" in cell and "amount" not in col_map:
                        col_map["amount"] = j
                if "amount" not in col_map:
                    for j, cell in enumerate(cells):
                        if "סכום עסקה" in cell:
                            col_map["amount"] = j
                            break
                break

        if header_row is None or "date" not in col_map or "amount" not in col_map:
            continue

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not any(row):
                continue
            date_cell = row[col_map["date"]] if len(row) > col_map["date"] else None
            # Footer rows have 'סך הכל' in col A and an amount-only row below it.
            if _norm_cell(date_cell) == "סך הכל":
                break

            try:
                desc_cell = row[col_map["description"]] if "description" in col_map and len(row) > col_map["description"] else None
                desc = _norm_cell(desc_cell) or "לא ידוע"
                amount_raw = row[col_map["amount"]] if len(row) > col_map["amount"] else None
                if amount_raw is None:
                    continue
                amount = float(str(amount_raw).replace(",", "").replace("₪", "").strip())
                # Keep negatives — refunds net out against the corresponding
                # monthly debit in Hapoalim (same reason as Cal).
                if amount == 0 or abs(amount) > 100000:
                    continue
                date = _format_excel_date(date_cell)
                if not date or date == "None":
                    continue
                transactions.append({
                    "date": date,
                    "description": normalize_description(desc),
                    "amount": amount,
                    "source": "מקס",
                    "type": "expense",
                })
            except (ValueError, TypeError, IndexError, AttributeError) as e:
                logger.debug(f"Skipped Max row: {e}")
                continue

    return transactions
